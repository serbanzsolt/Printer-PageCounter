import win32print
import openpyxl

# Connect to the network printer
# printer_name = "\\server\printer"
printer_name = "\\\\nfo-fs01\INFO Xerox VersaLink B7035"
printer_handle = win32print.OpenPrinter(printer_name)
printer_info = win32print.GetPrinter(printer_handle)

# Get the number of pages printed
pages_printed = printer_info['pdevmode'].dmFields & win32print.DM_PAGES

# Create an Excel file and write the results
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "Printer Name"
ws.cell(row=1, column=2).value = "Pages Printed"
ws.cell(row=2, column=1).value = printer_name
ws.cell(row=2, column=2).value = pages_printed
wb.save("printer_results.xlsx")
