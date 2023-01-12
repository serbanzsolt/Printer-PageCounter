import win32print
import openpyxl

# Connect to the network printer
printer_name = "\\\\nfo-fs01\BP2 RICOH 4510 IP205" # ensure the format of printer_name
printer_handle = win32print.OpenPrinter(printer_name)
print(printer_handle)
printer_info = win32print.GetPrinter(printer_handle)
print(printer_info)
print(printer_info[12])
pdevmode = printer_info[12] # Get pdevmode value

if pdevmode is None:
    print("Printer not found or disconnected")
else:
    pages_printed = pdevmode.dmFields & win32print.DM_PAGES

    # Create an Excel file and write the results
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Printer Name"
    ws.cell(row=1, column=2).value = "Pages Printed"
    ws.cell(row=2, column=1).value = printer_name
    ws.cell(row=2, column=2).value = pages_printed
    wb.save("printer_results.xlsx")
